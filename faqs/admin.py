from django.contrib import admin
from openpyxl import load_workbook
from django.utils.html import format_html
from .models import Faq, Answer, Step, Department, ResponseType, Event, Slide

# Register your models here.

@admin.register(Faq)
class FaqAdmin(admin.ModelAdmin):
    list_display = ('id', 'question', 'response_type', 'category', 'created_at', 'keywords', 'is_visible')
    search_fields = ('question', 'keywords')
    filter_horizontal = ('answers',)
    list_editable = ('is_visible',)  # Esto permite cambiarlo directamente desde la lista

    def get_answers(self, obj):
        return ", ".join([answer.answer_text for answer in obj.answers.all()])
    get_answers.short_description = 'answers'

@admin.register(Answer)
class AnswerAdmin(admin.ModelAdmin):
    list_display = ( 'id', 'title','node_type', 'answer_text', 'relevance', 'display_image', 'template', 'is_visible')
    search_fields = ('title','answer_text', 'steps')
    list_editable = ('is_visible',)  # Esto permite cambiarlo directamente desde la lista

    def display_image(self, obj):
        if obj.image:
            return format_html(f'<img src="{obj.image.url}" width="50" height="50" />')
        return "No Image"
    display_image.short_description = 'Image'

@admin.register(Step)
class StepAdmin(admin.ModelAdmin):
    list_display = ( 'id', 'number', 'text', 'answer', 'excel_file_preview')
    search_fields = ('text', 'answer')

    def excel_file_preview(self, obj):
        if obj.excel_file:
            return f"Ver contenido: {obj.excel_file.name}"
        return "No hay archivo"

    excel_file_preview.short_description = "Archivo Excel"

    def save_model(self, request, obj, form, change):
        # Procesar archivo Excel solo si está presente
        if obj.excel_file:
            try:
                workbook = load_workbook(obj.excel_file)
                sheet = workbook.active

                # Procesar filas del Excel
                excel_content = []
                for row in sheet.iter_rows(min_row=2, values_only=True):  # Salta encabezados
                    excel_content.append(row)
                    obj.text += f"{row}\n"  # Agrega al campo texto para propósitos de visualización

                # Opcional: Guarda el contenido del Excel como texto
                obj.text = "\n".join([str(row) for row in excel_content])

            except Exception as e:
                self.message_user(request, f"Error procesando archivo Excel: {e}", level='error')

        super().save_model(request, obj, form, change)

@admin.register(Department)
class CategoryAdmin(admin.ModelAdmin):
    list_display = ( 'id', 'name', 'description')
    search_fields = ('name', 'description')

@admin.register(ResponseType)
class ResponseTypeAdmin(admin.ModelAdmin):
    list_display = ('id', 'type_name', 'description')
    search_fields = ('type_name', 'description')

@admin.register(Event)
class EventoAdmin(admin.ModelAdmin):
    list_display = ( 'id', 'title', 'start_date', 'end_date', 'event_type')
    list_filter = ('event_type', 'start_date', 'county')
    search_fields = ('title', 'description', 'address', 'hospital', 'county')

@admin.register(Slide)
class SlideAdmin(admin.ModelAdmin):
    list_display = ( 'id', 'faq', 'question', 'left', 'right', 'up', 'down')
    search_fields = ('faq', 'question')