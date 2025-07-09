from django.db import models
from django.conf import settings
from rest_framework import serializers
from .models import Answer, AnswerConnection, Faq, Event, Step, Slide, Department
from django.db.models import Q # Importa Q para las condiciones de filtrado


class DepartmentSerializer(serializers.ModelSerializer):
    class Meta:
        model = Department
        fields = ['id', 'name'] # Puedes añadir 'description' si es útil

class AnswerConnectionSerializer(serializers.ModelSerializer):
    class Meta:
        model = AnswerConnection
        fields = ['to_answer', 'condition']

class StepSerializer(serializers.ModelSerializer):
    image_url = serializers.SerializerMethodField()
    excel_content = serializers.SerializerMethodField()

    class Meta:
        model = Step
        fields = ['number', 'text', 'image_url', 'excel_content']

    def get_excel_content(self, obj):
        if not obj.excel_file:
            return None
        try:
            from openpyxl import load_workbook
            workbook = load_workbook(obj.excel_file)
            sheet = workbook.active
            content = []
            for row in sheet.iter_rows(min_row=2, values_only=True):  # Salta encabezados
                content.append(row)
            return content
        except Exception as e:
            return str(e)        

    def get_image_url(self, obj):
        if obj.image:  # Asegúrate de que la imagen existe
            try:
                return obj.image.url
            except Exception as e:
                print(f"Error al obtener la URL de la imagen: {e}")
                return None
        return None

class AnswerSerializer(serializers.ModelSerializer):
    steps = StepSerializer(many=True)  # Incluye todos los datos de Step
    connections = AnswerConnectionSerializer(source='from_connections', many=True, required=False)
    image_url = serializers.SerializerMethodField()

    class Meta:
        model = Answer
        fields = '__all__'

    def get_image_url(self, obj):
        if obj.image:
            # Retorna la URL directa de Cloudinary
            return obj.image.url
        return None

class SlideSerializer(serializers.ModelSerializer):
    class Meta:
        model = Slide
        fields = ['id', 'faq', 'question', 'left', 'right', 'up', 'down']

class FaqSerializer(serializers.ModelSerializer):
    answers = serializers.SerializerMethodField() 
    slides = SlideSerializer(many=True, read_only=True)
    popularity = serializers.SerializerMethodField()
    response_type = serializers.SerializerMethodField()  
    category = serializers.SerializerMethodField()
    department = serializers.SerializerMethodField()

    class Meta:
        model = Faq
        fields = '__all__'

    def get_popularity(self, obj):
        # Calcula la popularidad en base a la relevancia acumulada de las respuestas
        return obj.answers.aggregate(total_relevance=models.Sum('relevance')).get('total_relevance') or 0
    
    def get_response_type(self, obj):
        return obj.response_type.type_name if obj.response_type else 'unknown'
    
    def get_department(self, obj):
        return obj.department.name if obj.department else 'unknown'

    def get_category(self, obj):
        print(f"Queue Type: {obj.category}")  # Para ver si se está llamando correctamente
        return obj.get_category_display()
    def get_answers(self, obj):
        query = self.context.get('query', '').strip()
        
        if query:
            # Filtra las Answers de esta FAQ que coinciden con el query
            # Asegúrate de que `keywords` sea un campo en tu modelo Answer que puedas buscar con `icontains`
            # (Si keywords es un ArrayField, icontains funciona para buscar un valor en el array)
            filtered_answers = obj.answers.filter(
                Q(answer_text__icontains=query) |
                Q(keywords__icontains=query) |
                Q(title__icontains=query) # También puedes considerar buscar en el título de la Answer
            ).distinct() # Usa distinct() por si una Answer tiene múltiples coincidencias con el mismo query
            return AnswerSerializer(filtered_answers, many=True).data
        else:
            # Si no hay query en el contexto, significa que esta FAQ no se encontró por una búsqueda de Answers
            # (o el endpoint no usó un query). En este caso, podemos devolver todas las Answers
            # o una lista vacía, dependiendo de tu lógica de negocio.
            # Para el contexto de 'search_faqs', una FAQ solo llega aquí si ya hubo una coincidencia.
            # Sin embargo, si FaqSerializer se usa en otro lugar sin un query,
            # esto aseguraría que se serialicen todas las respuestas asociadas a la FAQ.
            return AnswerSerializer(obj.answers.all(), many=True).data
    
class EventSerializer(serializers.ModelSerializer):
    class Meta:
        model = Event
        fields = '__all__'

